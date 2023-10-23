from openpyxl import Workbook, load_workbook

#grab the workbook from local file
wb = load_workbook('transactions.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['D1'] = 'corrected'
ws['D2'] = str(round(float(str(ws['C2'].value)[0:4].replace(',','.'))*0.9, 2)) + ' €'
ws['D3'] = str(round(float(str(ws['C3'].value)[0:4].replace(',','.'))*0.9, 2)) + ' €'
ws['D4'] = str(round(float(str(ws['C4'].value)[0:4].replace(',','.'))*0.9, 2)) + ' €'

wb.save('transactions.xlsx')